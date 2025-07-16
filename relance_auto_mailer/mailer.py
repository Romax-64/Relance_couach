# -*- coding: utf-8 -*-
"""
Created on Tue Jul  8 13:02:35 2025

@author: roman
"""

import logging
from pathlib import Path
from openpyxl import load_workbook, Workbook
import smtplib
from email.message import EmailMessage
from email.utils import make_msgid
from jinja2 import Environment, FileSystemLoader, select_autoescape

from .config import settings
from .utils import extract_valid_emails, get_today_str, generate_memory_filename

logger = logging.getLogger(__name__)

# Définition des chemins de templates et ressources
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR
RESOURCES_DIR = BASE_DIR.parent / "resources"

# Initialisation de l'environnement Jinja2
env = Environment(
    loader=FileSystemLoader(str(TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml", "tpl"])
)
# Chargement des templates
txt_tpl = env.get_template("relance_txt.tpl")  # Template texte brut
html_tpl = env.get_template("relance_html.tpl")  # Template HTML
sig_tpl = env.get_template("signature.tpl")  # Template signature

def envoyer_mails(fichier_relance: Path, sender_email: str, sender_password: str) -> Path:
    # Vérification de l'existence du fichier de relance
    relance_path = Path(fichier_relance)
    if not relance_path.exists():
        raise FileNotFoundError(f"Fichier de relance introuvable : {relance_path}")

    # Lecture du classeur Excel et sélection de la feuille active
    wb = load_workbook(relance_path, data_only=True)
    sheet = wb.active

    # Regroupement des commandes par fournisseur
    fournisseur_map = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extraction des colonnes pertinentes
        fournisseur = row[1]       # code fournisseur
        num_achat = row[2]         # numéro d'achat
        fourniture = row[4]        # libellé de l'article
        date_promise = row[7]      # date promise (JJ/MM/AAAA)
        mail_str = row[8]          # chaîne des e-mails
        code_pays = row[10]

        # Clé de regroupement: (fournisseur, liste d'e-mails)
        key = (fournisseur, mail_str)

        # Structure d'une commande pour le template
        cmd = {"num_achat": num_achat, "fourniture": fourniture, "date_promise": date_promise, "pays" : code_pays}
        fournisseur_map.setdefault(key, []).append(cmd)

    # Connexion au serveur SMTP
    smtp = smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT)
    smtp.starttls()  # Activation du chiffrement TLS
    smtp.login(sender_email, sender_password)

    # Préparation du mémo
    memoire_records = []
    date_relance = get_today_str()  # Date du jour formatée

    # Parcours de chaque fournisseur et envoi d'un e-mail
    for (fournisseur, mail_str), commandes in fournisseur_map.items():
        # Validation des adresses e-mail
        to_addrs = extract_valid_emails(mail_str)
        if not to_addrs:
            logger.warning(f"Aucune adresse valide pour {fournisseur}")
            continue

        # Détermination de la langue
        is_fr = settings.DEFAULT_LANGUAGE.upper() == 'FR'

        # Création d'un CID pour l'image logo
        logo_cid = make_msgid()[1:-1]

        # Contexte pour le rendu des templates
        context = {
            'commandes': commandes,
            'fournisseur': fournisseur,
            'date_relance': date_relance,
            'is_fr': is_fr,
            'logo_cid': logo_cid,
            'sender_name': settings.SENDER_NAME,
            'settings': settings
        }

        # Rendu des contenus texte et HTML
        text_body = txt_tpl.render(context)
        html_body = html_tpl.render(context) + sig_tpl.render(context)

        # Construction de l'e-mail
        msg = EmailMessage()
        msg['From'] = sender_email
        msg['To'] = ", ".join(to_addrs)
        msg['Subject'] = settings.SUBJECT_FR if is_fr else settings.SUBJECT_EN
        msg.set_content(text_body)
        msg.add_alternative(html_body, subtype='html')

        # Inclusion du logo via CID
        logo_path = RESOURCES_DIR / settings.LOGO_FILENAME
        if logo_path.exists():
            logo_data = logo_path.read_bytes()
            msg.get_payload()[1].add_related(
                logo_data,
                maintype='image', subtype='png', cid=f"<{logo_cid}>"
            )

        # Envoi du message
        #smtp.send_message(msg)
        logger.info(f"Email envoyé à {to_addrs} pour {fournisseur}")

        # Ajout des commandes au mémo
        for cmd in commandes:
            memoire_records.append([
                fournisseur,
                cmd['num_achat'],
                cmd['fourniture'],
                cmd['date_promise'],
                date_relance
            ])

    # Fermeture de la connexion SMTP
    smtp.quit()

    # Création du fichier de mémo des relances
    wb_out = Workbook()
    sheet_out = wb_out.active
    sheet_out.title = f"Memoire_{date_relance.replace('/', '_')}"
    sheet_out.append(['fournisseur', 'n° achat', 'fourniture', 'date promise', 'date relance'])
    for rec in memoire_records:
        sheet_out.append(rec)

    # Sauvegarde et retour du chemin du fichier mémo
    mem_file = generate_memory_filename()
    wb_out.save(mem_file)
    return mem_file
