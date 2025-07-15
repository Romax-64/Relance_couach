# -*- coding: utf-8 -*-
"""
Created on Tue Jul  8 13:02:35 2025

@author: roman
"""
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Tuple, Union
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)

# Définition du répertoire de base et du fichier fournisseurs fixe
BASE_DIR = Path(__file__).resolve().parent.parent
FOURNISSEURS_FILE = BASE_DIR / "resources" / "donnees_fournisseurs.xlsx"

def extract_data(
    fichier_commandes: Union[str, Path],
    fichier_fournisseurs: Union[str, Path] = FOURNISSEURS_FILE,
    fichier_sortie: Union[str, Path] = "relance_fournisseurs.xlsx"
) -> Tuple[int, Path]:
    # Conversion des chemins en objets Path
    fich_cmd = Path(fichier_commandes)
    fich_fou = Path(fichier_fournisseurs)
    sortie = Path(fichier_sortie)

    # Vérification de l'existence des fichiers
    if not fich_cmd.exists():
        logger.error(f"Fichier commandes introuvable : {fich_cmd}")
        raise FileNotFoundError(f"Fichier commandes introuvable : {fich_cmd}")
    if not fich_fou.exists():
        logger.error(f"Fichier fournisseurs introuvable : {fich_fou}")
        raise FileNotFoundError(f"Fichier fournisseurs introuvable : {fich_fou}")

    # Chargement et sélection de la feuille Retards ou feuille active si absente
    wb_cmd = load_workbook(fich_cmd, data_only=True)
    try:
        sheet_cmd = wb_cmd["Retards"]  # feuille nommée 'Retards'
    except KeyError:
        sheet_cmd = wb_cmd.active      # fallback vers la feuille active
        logger.warning(f"Feuille 'Retards' non trouvée, utilisation de '{sheet_cmd.title}'")

    # Chargement du fichier fournisseurs et lecture de la feuille principale
    wb_fou = load_workbook(fich_fou, data_only=True)
    sheet_fou = wb_fou.active

    # Construction d'un dictionnaire code_fournisseur -> informations contact
    contacts = {}
    for row in sheet_fou.iter_rows(min_row=2, values_only=True):
        code = row[0]                             # code fournisseur
        telephone = row[11] if len(row) > 11 else None  # numéro de téléphone
        mail = row[12] if len(row) > 12 else None        # adresse e-mail
        code_pays = row[13] if len(row) > 13 else None   # code pays
        contacts[code] = {
            "mail": mail or "",               # chaîne d'e-mails
            "telephone": telephone or "",
            "code_pays": code_pays or ""
        }

    # Extraction des commandes en retard et enrichissement des contacts
    records = []
    for idx, row in enumerate(sheet_cmd.iter_rows(min_row=2), start=1):
        code_fou    = row[0].value  # code fournisseur
        num_achat   = row[4].value  # numéro d'achat
        numeroligne = row[5].value  # numéro de ligne
        fourniture  = row[6].value  # nom de la fourniture
        designation = row[7].value  # désignation
        code_article= row[8].value  # code article
        date_prom   = row[14].value # date promise

        # Formatage de la date en JJ/MM/AAAA
        if isinstance(date_prom, (datetime, date)):
            date_str = date_prom.strftime("%d/%m/%Y")
        else:
            date_str = ""

        # Récupération des infos de contact pour ce fournisseur
        contact = contacts.get(code_fou, {"mail": "", "telephone": "", "code_pays": ""})

        # Assemblage du dictionnaire de données
        records.append({
            "ref": idx,
            "fournisseur": code_fou,
            "n° achat": num_achat,
            "n° ligne": numeroligne,
            "fourniture": fourniture,
            "désignation": designation,
            "code article": code_article,
            "date promise": date_str,
            "mail": contact["mail"],
            "téléphone": contact["telephone"],
            "code pays": contact["code_pays"]
        })

    # Tri alphabétique des enregistrements par fournisseur
    records.sort(key=lambda r: (r["fournisseur"] or "").upper())

    # Création et peuplement du fichier Excel de sortie
    wb_out = Workbook()
    sheet_out = wb_out.active
    sheet_out.title = "Relance fournisseurs"

    headers = [
        "ref", "fournisseur", "n° achat", "n° ligne",
        "fourniture", "désignation", "code article",
        "date promise", "mail", "téléphone", "code pays"
    ]
    sheet_out.append(headers)

    for rec in records:
        sheet_out.append([rec[h] for h in headers])

    # Sauvegarde du fichier et log de confirmation
    wb_out.save(sortie)
    logger.info(f"Fichier de relance généré : {sortie} ({len(records)} lignes)")

    return len(records), sortie

